#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Paso5_v2.py — Scorecards v2 (robusto) + Excel FULL + Beautify

Genera:
- Scorecards_v2.xlsx            → “Scorecards” + pestañas por sector
- Scorecards_v2_full.xlsx       → Main (valores ideales), Thresholds por sector, Detalle y pestañas por sector
"""

import argparse, math, os, re, sys
from typing import Dict, Any, Optional, Tuple, List

import numpy as np
import pandas as pd

# --------- Beautify de Excel (openpyxl) ---------
def beautify_excel(xlsx_path: str):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        print("Nota: para formatear Excel instalá 'openpyxl' (pip install openpyxl).", file=sys.stderr)
        return

    wb = load_workbook(xlsx_path)
    thin = Side(border_style="thin", color="DDDDDD")
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    percent_cols = {"WACC","ROIC_NOA","DCF_Upside","rf","mrp","tax_rate","g5","g_terminal"}
    pp_cols = {"Spread_pp"}
    money_cols = {"price","DCF_Fair"}
    int_cols = {"shares"}

    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        # header
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws.freeze_panes = "A2"

        headers = [cell.value for cell in ws[1]]

        # formatos numéricos
        for idx, name in enumerate(headers, start=1):
            for r in range(2, ws.max_row+1):
                cell = ws.cell(row=r, column=idx)
                if name in percent_cols:
                    cell.number_format = "0.00%"
                elif name in pp_cols:
                    cell.number_format = "0.00"
                elif name in money_cols:
                    cell.number_format = "#,##0.00"
                elif name in int_cols:
                    cell.number_format = "#,##0"

        # colores por Decision
        if "Decision" in headers:
            dcol = headers.index("Decision") + 1
            for r in range(2, ws.max_row+1):
                v = str(ws.cell(row=r, column=dcol).value or "").upper()
                color = None
                if v == "BUY":   color = "C6EFCE"
                elif v == "HOLD": color = "FFF2CC"
                elif v == "AVOID": color = "F8CBAD"
                if color:
                    ws.cell(row=r, column=dcol).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # auto-ancho
        for col in range(1, ws.max_column+1):
            letter = get_column_letter(col)
            maxlen = 10
            for r in range(1, min(ws.max_row, 500)+1):
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[letter].width = min(maxlen + 2, 40)

    wb.save(xlsx_path)

# -------------- deps online --------------
try:
    import yfinance as yf
except Exception:
    yf = None

# -------------- helpers ------------------
def clamp(x, lo, hi): return max(lo, min(hi, x))
def safe_div(a, b, default=np.nan):
    try:
        if b == 0 or pd.isna(a) or pd.isna(b): return default
        return a/b
    except Exception:
        return default

# -------------- config -------------------
DEFAULT_CFG = {
  "risk_free": {"USD":0.045,"EUR":0.030,"BRL":0.100,"ARS":0.200,"default":0.045},
  "mrp": {"USA":0.050,"Brazil":0.070,"Argentina":0.120,"default":0.055},
  "industry_unlevered_beta": {
      "Technology":1.00,"Consumer Cyclical":1.10,"Consumer Defensive":0.75,"Industrials":0.95,
      "Energy":1.15,"Utilities":0.60,"Healthcare":0.85,"Financial Services":0.70,"Real Estate":0.65,
      "Communication Services":0.95,"Basic Materials":1.05,"default":0.90
  },
  "terminal_growth": {"USA":0.020,"Brazil":0.025,"Argentina":0.030,"default":0.020},
  "wacc_floor":0.06, "wacc_cap":0.20, "default_tax_rate":0.25
}

def load_yaml(path:str)->dict:
    if not path: return DEFAULT_CFG
    try:
        import yaml
    except Exception:
        return DEFAULT_CFG
    try:
        with open(path,"r",encoding="utf-8") as f:
            raw = yaml.safe_load(f) or {}
        cfg = DEFAULT_CFG.copy()
        for k,v in raw.items():
            if isinstance(v, dict) and k in cfg:
                cfg[k] = {**cfg[k], **v}
            else:
                cfg[k] = v
        return cfg
    except Exception:
        return DEFAULT_CFG

def read_overrides(path:str)->Optional[pd.DataFrame]:
    try:
        if path and os.path.exists(path):
            return pd.read_csv(path, comment="#")
    except Exception:
        pass
    return None

# -------------- fetch & sector -----------
def fetch_company(ticker: str) -> Dict[str, Any]:
    if yf is None:
        raise RuntimeError("yfinance no disponible. Instalá dependencias y usá internet.")
    tk = yf.Ticker(ticker)

    info = tk.info or {}

    # ¡NO usar 'or' con DataFrames!
    bs = tk.balance_sheet
    if not isinstance(bs, pd.DataFrame) or bs.empty:
        bs = pd.DataFrame()

    is_ = tk.financials
    if not isinstance(is_, pd.DataFrame) or is_.empty:
        is_ = pd.DataFrame()

    cf = tk.cashflow
    if not isinstance(cf, pd.DataFrame) or cf.empty:
        cf = pd.DataFrame()

    if not bs.empty: bs = bs.fillna(0)
    if not is_.empty: is_ = is_.fillna(0)
    if not cf.empty: cf = cf.fillna(0)

    return dict(info=info, bs=bs, is_=is_, cf=cf)

def is_financial_sector(info:Dict[str,Any])->bool:
    sector = (info or {}).get("sector") or ""
    return "Financial" in sector or "Insurance" in sector

# -------------- métricas -----------------
def estimate_tax_rate(is_:pd.DataFrame, default_rate:float)->float:
    try:
        col = is_.columns[0]
        tax = float(is_.loc["Income Tax Expense", col])
        ebt = float(is_.loc["Pretax Income", col])
        tr = clamp(safe_div(tax, ebt, default_rate), 0.0, 0.35)
        return tr if not math.isnan(tr) else default_rate
    except Exception:
        return default_rate

def ebit_last(is_:pd.DataFrame)->float:
    for k in ["Ebit","EBIT","Operating Income"]:
        if k in is_.index:
            try: return float(is_.loc[k, is_.columns[0]])
            except Exception: pass
    try:
        col=is_.columns[0]
        gp=float(is_.loc["Gross Profit", col])
        sga=float(is_.loc.get("Selling General Administrative", pd.Series({col:0}))[col])
        rd=float(is_.loc.get("Research Development", pd.Series({col:0}))[col])
        return gp - sga - rd
    except Exception:
        return np.nan

def compute_nopat(is_:pd.DataFrame, tax:float)->float:
    e=ebit_last(is_)
    return np.nan if pd.isna(e) else e*(1-tax)

def compute_noa(bs:pd.DataFrame)->float:
    try:
        col=bs.columns[0]
        ta=float(bs.loc["Total Assets", col])
        cash=float(bs.loc.get("Cash And Cash Equivalents", pd.Series({col:0}))[col])
        sti=float(bs.loc.get("Short Term Investments", pd.Series({col:0}))[col])
        st_debt=float(bs.loc.get("Short Long Term Debt", pd.Series({col:0}))[col])
        lt_debt=float(bs.loc.get("Long Term Debt", pd.Series({col:0}))[col])
        tl=float(bs.loc["Total Liab", col])
        op_liab=max(tl-(st_debt+lt_debt),0.0)
        op_assets=max(ta-(cash+sti),0.0)
        return max(op_assets - op_liab, 0.0)
    except Exception:
        return np.nan

def compute_roic(nopat:float, noa_now:float, noa_prev:Optional[float])->float:
    if pd.isna(nopat) or pd.isna(noa_now): return np.nan
    noa_avg=np.nanmean([noa_now, noa_prev if (noa_prev is not None and not pd.isna(noa_prev)) else noa_now])
    return safe_div(nopat, noa_avg, np.nan)

def market_weights(info:Dict[str,Any], bs:pd.DataFrame)->Tuple[float,float]:
    try:
        price=info.get("currentPrice") or info.get("previousClose") or np.nan
        shares=info.get("sharesOutstanding") or np.nan
        mcap=price*shares
    except Exception: mcap=np.nan
    try:
        col=bs.columns[0]
        st=float(bs.loc.get("Short Long Term Debt", pd.Series({col:0}))[col])
        lt=float(bs.loc.get("Long Term Debt", pd.Series({col:0}))[col])
        cash=float(bs.loc.get("Cash And Cash Equivalents", pd.Series({col:0}))[col])
        debt=max(st+lt-cash,0.0)
    except Exception: debt=np.nan
    if pd.isna(mcap) or pd.isna(debt) or (mcap+debt)==0: return np.nan, np.nan
    tot=mcap+debt
    return mcap/tot, debt/tot

def book_weights(bs:pd.DataFrame)->Tuple[float,float]:
    try:
        col=bs.columns[0]
        eq=float(bs.loc.get("Total Stockholder Equity", pd.Series({col:np.nan}))[col])
        st=float(bs.loc.get("Short Long Term Debt", pd.Series({col:0}))[col])
        lt=float(bs.loc.get("Long Term Debt", pd.Series({col:0}))[col])
        debt=max(st+lt,0.0); tot=max(debt+max(eq,0.0),0.0)
        if tot==0: return np.nan,np.nan
        return max(eq,0.0)/tot, debt/tot
    except Exception:
        return np.nan,np.nan

def cost_of_debt(is_:pd.DataFrame, bs:pd.DataFrame, rf:float, spread_override:Optional[float])->float:
    if spread_override is not None and not pd.isna(spread_override):
        return max(rf+spread_override, rf)
    try:
        col=is_.columns[0]
        interest=float(is_.loc.get("Interest Expense", pd.Series({col:np.nan}))[col])
    except Exception: interest=np.nan
    try:
        colb=bs.columns[0]
        st=float(bs.loc.get("Short Long Term Debt", pd.Series({colb:0}))[colb])
        lt=float(bs.loc.get("Long Term Debt", pd.Series({colb:0}))[colb])
        debt=max(st+lt,0.0)
    except Exception: debt=np.nan
    kd=safe_div(abs(interest), debt if debt>0 else np.nan, np.nan)
    if pd.isna(kd): kd=rf+0.02
    return max(kd, rf)

def relever_beta(beta_u:float, we:float, wd:float, tax:float)->float:
    if pd.isna(we) or pd.isna(wd) or we<=0: we,wd=0.6,0.4
    return beta_u*(1+(1-tax)*wd/we)

def compute_wacc(cfg:dict, info:Dict[str,Any], is_:pd.DataFrame, bs:pd.DataFrame, ov:Dict[str,Any])->Dict[str,float]:
    country = ov.get("country") or info.get("country") or "USA"
    currency = ov.get("currency") or info.get("currency") or "USD"
    rf = ov.get("rf", np.nan);  rf = rf if not pd.isna(rf) else cfg["risk_free"].get(currency, cfg["risk_free"]["default"])
    mrp = ov.get("mrp", np.nan); mrp = mrp if not pd.isna(mrp) else cfg["mrp"].get(country, cfg["mrp"]["default"])
    tax = ov.get("tax_rate", np.nan); tax = tax if not pd.isna(tax) else estimate_tax_rate(is_, cfg["default_tax_rate"])
    sector = info.get("sector") or "default"
    bu = ov.get("industry_unlevered_beta", np.nan); bu = bu if not pd.isna(bu) else cfg["industry_unlevered_beta"].get(sector, cfg["industry_unlevered_beta"]["default"])
    we, wd = market_weights(info, bs)
    if pd.isna(we) or pd.isna(wd): we, wd = book_weights(bs)
    beta_l = relever_beta(bu, we, wd, tax)
    kd = cost_of_debt(is_, bs, rf, ov.get("debt_spread", None))
    ke = rf + beta_l*mrp
    wacc = clamp(we*ke + wd*kd*(1-tax), cfg["wacc_floor"], cfg["wacc_cap"])
    return dict(wacc=wacc, rf=rf, mrp=mrp, tax=tax, beta_u=bu, beta_l=beta_l, kd=kd, ke=ke,
                we=we, wd=wd, country=country, currency=currency)

def revenue_cagr_3y(is_:pd.DataFrame)->float:
    try:
        cols=list(is_.columns)
        if len(cols)<3: return np.nan
        rev0=float(is_.loc["Total Revenue", cols[0]])
        rev2=float(is_.loc["Total Revenue", cols[2]])
        return (rev0/rev2)**(1/2.0)-1
    except Exception:
        return np.nan

def last_fcf(cf:pd.DataFrame)->float:
    try:
        col=cf.columns[0]
        cfo=float(cf.loc.get("Total Cash From Operating Activities", pd.Series({col:np.nan}))[col])
        capex=float(cf.loc.get("Capital Expenditures", pd.Series({col:0}))[col])
        return cfo - abs(capex)
    except Exception:
        return np.nan

def dcf_fair_value(cfg:dict, info:Dict[str,Any], is_:pd.DataFrame, cf:pd.DataFrame, wacc:float, ov:Dict[str,Any])->Dict[str,float]:
    country=ov.get("country") or info.get("country") or "USA"
    g5 = ov.get("g5", np.nan);  g5 = g5 if not pd.isna(g5) else revenue_cagr_3y(is_)
    if pd.isna(g5): g5 = 0.04
    g5 = clamp(g5, -0.05, 0.10)
    gterm = ov.get("g_terminal", np.nan); gterm = gterm if not pd.isna(gterm) else cfg["terminal_growth"].get(country, cfg["terminal_growth"]["default"])
    gterm = clamp(gterm, 0.0, min(0.03, wacc-0.01))
    f0 = last_fcf(cf)
    if pd.isna(f0): return dict(fair_value=np.nan, upside=np.nan, g5=g5, gterm=gterm)
    f = [f0*((1+g5)**i) for i in range(1,6)]
    disc = [(1/((1+wacc)**i)) for i in range(1,6)]
    pv5 = sum(fi*di for fi,di in zip(f,disc))
    tv = f[-1]*(1+gterm)/(wacc-gterm)
    ev = pv5 + tv*disc[-1]
    price = info.get("currentPrice") or info.get("previousClose") or np.nan
    shares = info.get("sharesOutstanding") or np.nan
    if pd.isna(price) or pd.isna(shares):
        return dict(fair_value=np.nan, upside=np.nan, ev=ev, g5=g5, gterm=gterm)
    fair = ev/shares
    return dict(fair_value=fair, upside=(fair/price)-1, ev=ev, g5=g5, gterm=gterm)

def piotroski_f_score(is_:pd.DataFrame, bs:pd.DataFrame, cf:pd.DataFrame)->float:
    try:
        cols=list(is_.columns)
        if len(cols)<2: return np.nan
        c0,c1=cols[0], cols[1]
        ni0=float(is_.loc["Net Income", c0]); ni1=float(is_.loc["Net Income", c1])
        roa0=safe_div(ni0,float(bs.loc["Total Assets", c0])); roa1=safe_div(ni1,float(bs.loc["Total Assets", c1]))
        cfo0=float(cf.loc["Total Cash From Operating Activities", c0])
        ldt0=float(bs.loc.get("Long Term Debt", pd.Series({c0:0}))[c0]); ldt1=float(bs.loc.get("Long Term Debt", pd.Series({c1:0}))[c1])
        cr0=safe_div(float(bs.loc["Total Current Assets", c0]), float(bs.loc["Total Current Liabilities", c0]))
        cr1=safe_div(float(bs.loc["Total Current Assets", c1]), float(bs.loc["Total Current Liabilities", c1]))
        gm0=safe_div(float(is_.loc["Gross Profit", c0]), float(is_.loc["Total Revenue", c0]))
        gm1=safe_div(float(is_.loc["Gross Profit", c1]), float(is_.loc["Total Revenue", c1]))
        at0=safe_div(float(is_.loc["Total Revenue", c0]), float(bs.loc["Total Assets", c0]))
        at1=safe_div(float(is_.loc["Total Revenue", c1]), float(bs.loc["Total Assets", c1]))
        score=0
        score+= 1 if ni0>0 else 0
        score+= 1 if cfo0>0 else 0
        score+= 1 if roa0>roa1 else 0
        score+= 1 if cfo0>ni0 else 0
        score+= 1 if ldt0<=ldt1 else 0
        score+= 1 if cr0>cr1 else 0
        score+= 1 if gm0>gm1 else 0
        score+= 1 if at0>at1 else 0
        return float(score)
    except Exception:
        return np.nan

def altman_z_score(info:Dict[str,Any], is_:pd.DataFrame, bs:pd.DataFrame, mcap:Optional[float])->float:
    try:
        col=bs.columns[0]
        wc=float(bs.loc["Total Current Assets", col]) - float(bs.loc["Total Current Liabilities", col])
        re=float(bs.loc.get("Retained Earnings", pd.Series({col:0}))[col])
        ebit=ebit_last(is_)
        sales=float(is_.loc["Total Revenue", is_.columns[0]])
        ta=float(bs.loc["Total Assets", col])
        if pd.isna(mcap):
            price=info.get("currentPrice") or info.get("previousClose") or np.nan
            shares=info.get("sharesOutstanding") or np.nan
            mcap=price*shares if not (pd.isna(price) or pd.isna(shares)) else np.nan
        if pd.isna(mcap):
            be=float(bs.loc.get("Total Stockholder Equity", pd.Series({col:np.nan}))[col])
            mcap=max(be,0.0)
        x1=safe_div(wc,ta,0); x2=safe_div(re,ta,0); x3=safe_div(ebit,ta,0)
        x4=safe_div(mcap,float(bs.loc["Total Liab", col]),0); x5=safe_div(sales,ta,0)
        return 1.2*x1 + 1.4*x2 + 3.3*x3 + 0.6*x4 + 1.0*x5
    except Exception:
        return np.nan

# -------------- pipeline core ------------
ABS_SP, ABS_F, ABS_Z, ABS_UP = 2.0, 7, 3.0, 0.20

def analyze_tickers(tickers:List[str], cfg:dict, ov_df:Optional[pd.DataFrame])->pd.DataFrame:
    rows=[]
    for t in tickers:
        ov={}
        if ov_df is not None and not ov_df.empty:
            r=ov_df[ov_df["ticker"].astype(str).str.upper()==t.upper()]
            if not r.empty:
                ov={k:r.iloc[0][k] for k in r.columns if not (isinstance(r.iloc[0][k], float) and math.isnan(r.iloc[0][k]))}
        data=fetch_company(t)
        info,bs,is_,cf = data["info"],data["bs"],data["is_"],data["cf"]
        fin = is_financial_sector(info)
        tax = estimate_tax_rate(is_, DEFAULT_CFG["default_tax_rate"])
        wacc_pack = compute_wacc(cfg, info, is_, bs, ov)
        wacc = wacc_pack["wacc"]
        noa_now = compute_noa(bs)
        noa_prev = compute_noa(bs.iloc[:,[1]]) if isinstance(bs, pd.DataFrame) and bs.shape[1]>=2 else np.nan
        nopat = compute_nopat(is_, tax)
        roic = compute_roic(nopat, noa_now, noa_prev)
        price = info.get("currentPrice") or info.get("previousClose") or np.nan
        shares = info.get("sharesOutstanding") or np.nan
        mcap = price*shares if not (pd.isna(price) or pd.isna(shares)) else np.nan
        fscore = piotroski_f_score(is_, bs, cf)
        z = np.nan if fin else altman_z_score(info, is_, bs, mcap)
        fv = dict(fair_value=np.nan, upside=np.nan, g5=np.nan, gterm=np.nan, ev=np.nan) if fin else dcf_fair_value(cfg, info, is_, cf, wacc, ov)
        spread = (roic - wacc) if (not pd.isna(roic) and not pd.isna(wacc)) else np.nan
        rows.append({
            "ticker":t,"company":info.get("longName") or t,"sector":info.get("sector") or "Unknown",
            "country":wacc_pack["country"],"currency":wacc_pack["currency"],
            "price":price,"shares":shares,"market_cap":mcap,
            "WACC":wacc,"ROIC_NOA":roic,"Spread_pp": spread*100 if not pd.isna(spread) else np.nan,
            "F_Score":fscore,"Altman_Z":z,
            "DCF_Fair":fv.get("fair_value",np.nan),"DCF_Upside":fv.get("upside",np.nan),"DCF_EV":fv.get("ev",np.nan),
            "g5":fv.get("g5",np.nan),"g_terminal":fv.get("gterm",np.nan),
            "NOPAT":nopat,"NOA_now":noa_now,"NOA_prev":noa_prev,
            "rf":wacc_pack["rf"],"mrp":wacc_pack["mrp"],"tax_rate":wacc_pack["tax"],
            "beta_u":wacc_pack["beta_u"],"beta_l":wacc_pack["beta_l"],"ke":wacc_pack["ke"],"kd":wacc_pack["kd"],
            "wE":wacc_pack["we"],"wD":wacc_pack["wd"]
        })
    df=pd.DataFrame(rows)
    if df.empty: return df

    def decide(g: pd.DataFrame):
        if len(g) < 8:
            t_sp, t_f, t_z, t_up = ABS_SP, ABS_F, ABS_Z, ABS_UP
        else:
            t_sp = max(g["Spread_pp"].quantile(0.60), ABS_SP)
            t_f  = max(g["F_Score" ].quantile(0.60), ABS_F)
            t_z  = max(g["Altman_Z"].quantile(0.60), ABS_Z)
            t_up = max(g["DCF_Upside"].quantile(0.60), ABS_UP)

        out = []
        for _, r in g.iterrows():
            if pd.notna(r.get("Altman_Z")) and r["Altman_Z"] < 1.8:
                out.append("AVOID"); continue
            if pd.notna(r.get("F_Score")) and r["F_Score"] <= 3:
                out.append("AVOID"); continue

            s1 = (r["Spread_pp"]  >= t_sp) if pd.notna(r["Spread_pp"])  else False
            s2 = (r["F_Score"]    >= t_f ) if pd.notna(r["F_Score"])    else False
            s3 = (r["Altman_Z"]   >= t_z ) if pd.notna(r["Altman_Z"])   else False
            s4 = (r["DCF_Upside"] >= t_up) if pd.notna(r["DCF_Upside"]) else False
            sig = s1 + s2 + s3 + s4
            out.append("BUY" if sig >= 3 else ("HOLD" if sig == 2 else "AVOID"))
        g = g.copy()
        g["Decision"] = out
        return g

    return df.groupby("sector", group_keys=False).apply(decide)

# ------------------ exports ----------------
def export_excel(df:pd.DataFrame, path:str):
    with pd.ExcelWriter(path, engine="xlsxwriter") as w:
        df.to_excel(w, index=False, sheet_name="Scorecards")
        for sec, sub in df.groupby("sector"):
            sub.to_excel(w, index=False, sheet_name=sec[:31])

def build_thresholds_by_sector(df: pd.DataFrame)->pd.DataFrame:
    rows=[]
    for sec, g in df.groupby("sector"):
        n=len(g)
        if n<8:
            rule="ABS_ONLY"
            t_sp, t_f, t_z, t_up = ABS_SP, ABS_F, ABS_Z, ABS_UP
        else:
            rule="P60_AND_ABS"
            t_sp = max(g["Spread_pp"].quantile(0.60), ABS_SP)
            t_f  = max(g["F_Score" ].quantile(0.60), ABS_F)
            t_z  = max(g["Altman_Z"].quantile(0.60), ABS_Z)
            t_up = max(g["DCF_Upside"].quantile(0.60), ABS_UP)
        rows.append({
            "sector": sec, "n": n, "rule": rule,
            "Thr_Spread_pp": t_sp, "Thr_F_Score": t_f, "Thr_Altman_Z": t_z, "Thr_DCF_Upside": t_up
        })
    return pd.DataFrame(rows).sort_values(["rule","sector"]).reset_index(drop=True)

def export_full_excel(df:pd.DataFrame, path:str):
    base_cols = [
        "ticker","company","sector","country","currency","price","market_cap","shares",
        "Decision","Spread_pp","F_Score","Altman_Z","DCF_Upside","DCF_Fair",
        "WACC","ROIC_NOA","NOPAT","NOA_now","NOA_prev",
        "rf","mrp","tax_rate","beta_u","beta_l","ke","kd","wE","wD",
        "g5","g_terminal","DCF_EV"
    ]
    cols_order = [c for c in base_cols if c in df.columns] + [c for c in df.columns if c not in base_cols]
    thr = build_thresholds_by_sector(df)
    with pd.ExcelWriter(path, engine="xlsxwriter") as w:
        pd.DataFrame({
            "Métrica": ["Spread (pp)","Piotroski F-Score","Altman Z","Upside DCF"],
            "Ideal (piso absoluto)": [ABS_SP, ABS_F, ABS_Z, ABS_UP]
        }).to_excel(w, index=False, sheet_name="Main")
        thr.to_excel(w, index=False, sheet_name="Thresholds_sector")
        df[cols_order].to_excel(w, index=False, sheet_name="Detalle")
        for sec, sub in df.groupby("sector"):
            sub = sub.sort_values(["Decision","Spread_pp","F_Score"], ascending=[True, False, False], na_position="last")
            sub.to_excel(w, index=False, sheet_name=sec[:31])

# -------------- CLI -----------------------
def parse_tickers_arg(raw:Optional[List[str]])->List[str]:
    if not raw:
        txt = input("Ingresá tickers separados por coma o espacio (ej: AAPL MSFT MELI): ").strip()
    else:
        txt = " ".join(raw)
    items = re.split(r"[,\s]+", txt.strip())
    return [t for t in items if t]

def main():
    ap=argparse.ArgumentParser(description="Scorecards v2 (standalone)")
    ap.add_argument("--tickers","-t", nargs="*", help="AAPL MSFT o 'AAPL,MSFT'")
    ap.add_argument("--out", default="Scorecards_v2.xlsx")
    ap.add_argument("--full-out", default="Scorecards_v2_full.xlsx",
                    help="XLSX FULL: Main (ideales) + thresholds por sector + detalle completo")
    ap.add_argument("--config","-c", default="", help="finance_config.yaml (opcional)")
    ap.add_argument("--overrides","-o", default="", help="overrides.csv (opcional)")
    args=ap.parse_args()

    if yf is None:
        print("Error: yfinance no está disponible. Instalá dependencias e internet.", file=sys.stderr); sys.exit(2)

    cfg=load_yaml(args.config)
    ov=read_overrides(args.overrides)
    tickers=parse_tickers_arg(args.tickers)
    if not tickers: print("No hay tickers."); sys.exit(1)

    df=analyze_tickers(tickers,cfg,ov)
    if df is None or df.empty:
        print("No se pudo generar scorecard v2."); sys.exit(2)

    export_excel(df, args.out)
    export_full_excel(df, args.full_out)

    # beautify
    for p in [args.out, args.full_out]:
        try:
            beautify_excel(p)
            print(f"✓ Formato aplicado a {p}")
        except Exception as e:
            print(f"ADVERTENCIA: no pude formatear {p}: {e}", file=sys.stderr)

    cols=[c for c in ["ticker","Decision","Spread_pp","F_Score","Altman_Z","DCF_Upside","WACC","ROIC_NOA"] if c in df.columns]
    print(df[cols].to_string(index=False))
    print(f"✔ Scorecards v2 generado: {args.out}")
    print(f"✔ Scorecards v2 FULL generado: {args.full_out}")

if __name__=="__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
